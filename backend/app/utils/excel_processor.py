"""Excel 檔案處理工具。"""

import io
import os
from typing import Any

import openpyxl
import pandas as pd
import structlog

logger = structlog.get_logger()


def smart_read_excel(
    file_path_or_obj: str | os.PathLike[str] | io.BytesIO,
    header_rows: int = 1,
    sheet_name: int | str = 0,
    unmerge: bool = True,
) -> pd.DataFrame:
    """Enhanced Excel reader that handles merged cells and multi-row headers.

    Args:
        file_path_or_obj: Path to file or file-like object.
        header_rows: Number of rows that represent the header (for multi-level headers).
        sheet_name: Index or name of the sheet to read.
        unmerge: Whether to unmerge cells and fill values.

    Returns:
        pd.DataFrame: Cleaned dataframe.
    """
    try:
        if isinstance(file_path_or_obj, (str, os.PathLike)):
            wb = openpyxl.load_workbook(file_path_or_obj, data_only=True)
        else:
            content = (
                file_path_or_obj.read()
                if hasattr(file_path_or_obj, "read")
                else file_path_or_obj
            )
            wb = openpyxl.load_workbook(
                io.BytesIO(content) if isinstance(content, bytes) else content,
                data_only=True,
            )

        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]

        if unmerge:
            merged_cells_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_cells_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_value = sheet.cell(row=min_row, column=min_col).value

                sheet.unmerge_cells(str(merged_range))

                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        sheet.cell(row=r, column=c).value = top_left_value

        data: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        if not data:
            return pd.DataFrame()

        if header_rows > 1:
            header_data = data[:header_rows]
            body_data = data[header_rows:]

            flat_headers: list[str] = []
            num_cols = len(header_data[0]) if header_data else 0

            for col_idx in range(num_cols):
                parts: list[str] = []
                for row_idx in range(header_rows):
                    val = header_data[row_idx][col_idx]
                    if val is not None and str(val).strip():
                        parts.append(str(val).strip())

                seen_parts: list[str] = []
                for p in parts:
                    if p not in seen_parts:
                        seen_parts.append(p)

                header_name = "_".join(seen_parts)
                if not header_name:
                    header_name = f"Column_{col_idx + 1}"
                flat_headers.append(header_name)

            df = pd.DataFrame(body_data, columns=flat_headers)
        else:
            df = pd.DataFrame(data[1:], columns=data[0])

        df.columns = [
            str(c) if c is not None else f"Unnamed_{i}"
            for i, c in enumerate(df.columns)
        ]
        df.columns = [c.strip() for c in df.columns]
        df = df.dropna(how="all").dropna(axis=1, how="all")

        return df

    except Exception as e:
        logger.error(f"Error in smart_read_excel: {str(e)}")
        try:
            return pd.read_excel(
                file_path_or_obj,
                header=list(range(header_rows)) if header_rows > 1 else 0,
            )
        except Exception:
            return pd.DataFrame()
